"""
Write a per-run Excel log: a copy of the source sheet rows plus a Status column,
updated row-by-row as the runner processes each SKU. Saved incrementally so a
crash mid-run preserves partial progress.

Status values:
  success         — offer created (and patched / published)
  skipped         — SKU not in Allegro catalog
  multi_variant   — multiple variants returned, picked first
  error           — API failure (cookies, datadome, network, server)
"""

from datetime import datetime
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from sheets import SheetRow


HEADER = ["Row", "SKU", "Price", "Stock", "Status", "Offer ID", "Notes", "Timestamp"]

STATUS_FILLS = {
    "success": PatternFill("solid", fgColor="C6EFCE"),
    "skipped": PatternFill("solid", fgColor="FFEB9C"),
    "multi_variant": PatternFill("solid", fgColor="FFD8B0"),
    "error": PatternFill("solid", fgColor="FFC7CE"),
}
STATUS_FONTS = {
    "success": Font(color="006100"),
    "skipped": Font(color="9C5700"),
    "multi_variant": Font(color="C56500"),
    "error": Font(color="9C0006"),
}


class RunExcelWriter:
    """Writes & updates run_<id>.xlsx incrementally."""

    def __init__(self, path: Path, sheet_rows: Iterable[SheetRow]) -> None:
        self.path = path
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "Run"

        self.ws.append(HEADER)
        for cell in self.ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        self._row_index: dict[int, int] = {}
        for i, sr in enumerate(sheet_rows, start=2):
            self.ws.append([sr.row, sr.sku, sr.price, sr.stock, "", "", "", ""])
            self._row_index[sr.row] = i

        widths = [8, 22, 12, 10, 16, 18, 50, 22]
        for i, w in enumerate(widths, start=1):
            self.ws.column_dimensions[get_column_letter(i)].width = w

        self.save()

    def save(self) -> None:
        self.wb.save(self.path)

    def update(
        self,
        sheet_row: int,
        status: str,
        offer_id: str = "",
        notes: str = "",
    ) -> None:
        excel_row = self._row_index.get(sheet_row)
        if not excel_row:
            return
        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.ws.cell(row=excel_row, column=5, value=status)
        self.ws.cell(row=excel_row, column=6, value=offer_id)
        self.ws.cell(row=excel_row, column=7, value=notes)
        self.ws.cell(row=excel_row, column=8, value=ts)

        fill = STATUS_FILLS.get(status)
        font = STATUS_FONTS.get(status)
        if fill:
            self.ws.cell(row=excel_row, column=5).fill = fill
        if font:
            self.ws.cell(row=excel_row, column=5).font = font

        self.save()
